from tkinter import messagebox
from typing import Optional, Union, List, Dict, Any, Hashable

import pandas as pd
import openpyxl  # Добавляем для работы с цветами
import scripts.handlings.handling_json as handling_json


class ExcelReader:
    """
    Базовый класс для чтения Excel-файлов.
    Объединяет всю логику загрузки данных и общие операции.
    """
    
    def __init__(self, file_path: str, sheet_name: Optional[Union[str, int]] = None,
                 color_filter_column: Optional[str] = None,  # Новый параметр
                 track_sheet_origin: bool = False):           # Новый параметр
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.data: Optional[pd.DataFrame] = None
        self.columns_save: List[str] = []
        self.filtered_data: Optional[Dict[int, Dict[str, Any]]] = None
        self.color_filter_column = color_filter_column      # Сохраняем
        self.track_sheet_origin = track_sheet_origin        # Сохраняем
        self.sheet_origin: Optional[str] = None             # Для MultiSheetReader
        
        self.load_excel()
    
    def load_excel(self, file_path: Optional[str] = None, 
                   sheet_name: Optional[Union[str, int]] = None,
                   color_filter_column: Optional[str] = None) -> Optional[pd.DataFrame]:
        """
        Загружает данные из Excel файла.
        """
        path = file_path or self.file_path
        sheet = sheet_name if sheet_name is not None else self.sheet_name
        color_col = color_filter_column or self.color_filter_column
        
        try:
            # Сначала читаем через openpyxl для проверки цветов
            if color_col is not None:
                data = self._load_with_color_filter(path, sheet, color_col)
            else:
                data = pd.read_excel(path, sheet_name=sheet)
            
            # Если вернулся словарь (несколько листов), берём первый
            if isinstance(data, dict):
                sheet_names = list(data.keys())
                if sheet_names:
                    first_sheet = sheet_names[0]
                    print(f"Лист не указан. Загружаем первый лист: {first_sheet}")
                    data = data[first_sheet]
                else:
                    raise ValueError("Файл Excel не содержит листов.")
            
            if file_path is None and sheet_name is None:
                self.data = data
                print(f"Лист успешно загружен: {sheet if sheet else 'первый лист'}")
            
            return data
            
        except Exception as e:
            print(f"Ошибка при загрузке файла или листа: {e}")
            if file_path is None and sheet_name is None:
                self.data = None
            return None

    def _load_with_color_filter(self, path: str, sheet: Optional[Union[str, int]],
                                color_col: str) -> pd.DataFrame:
        """Загружает данные с цветной фильтрацией и сохраняет цвета."""
        wb = openpyxl.load_workbook(path, data_only=True)

        if sheet is None:
            ws = wb.active
            self.sheet_origin = ws.title
        elif isinstance(sheet, int):
            ws = wb.worksheets[sheet]
            self.sheet_origin = ws.title
        else:
            ws = wb[sheet]
            self.sheet_origin = sheet

        header_row = 1
        headers = [cell.value for cell in ws[header_row]]

        try:
            col_idx = headers.index(color_col) + 1
        except ValueError:
            wb.close()
            raise ValueError(f"Колонка '{color_col}' не найдена")

        colored_rows = []
        row_colors = []  # Сохраняем цвета

        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            color = None
            has_color = False

            if cell.fill and cell.fill.fgColor:
                rgb = cell.fill.fgColor.rgb
                if rgb and rgb not in ('00000000', 'FFFFFFFF', None):
                    color = str(rgb)
                    has_color = True

            if has_color:
                colored_rows.append(row_idx - header_row - 1)
                row_colors.append(color)

        wb.close()

        df = pd.read_excel(path, sheet_name=sheet)
        filtered_df = df.iloc[colored_rows].reset_index(drop=True)

        # Добавляем служебные колонки
        if self.track_sheet_origin:
            filtered_df['__sheet_origin__'] = self.sheet_origin

        # Добавляем цвета как отдельную колонку
        filtered_df['__color_D__'] = row_colors

        return filtered_df
    
    def get_dict_all_data(self) -> Dict[int, Dict[str, Any]]:
        """
        Возвращает весь словарь данных из self.data.
        """
        if self.data is None:
            print("Данные не загружены.")
            return {}
        
        return {index: row.to_dict() for index, row in self.data.iterrows()}
    
    def get_headers(self) -> List[str]:
        """Возвращает список заголовков колонок."""
        if self.data is None:
            return []
        return list(self.data.columns)
    
    def filter_and_save_columns(self, columns_to_save: Union[str, List[str], tuple]):
        """
        Сохраняет значения из указанных столбцов в словарь.
        """
        if self.data is None:
            print("Данные не загружены.")
            return
        
        if isinstance(columns_to_save, str):
            columns_to_save = [columns_to_save]
        elif not isinstance(columns_to_save, list):
            columns_to_save = list(columns_to_save)
        
        self.columns_save = columns_to_save
        self.filtered_data = {}
        
        for index, row in self.data.iterrows():
            self.filtered_data[index] = {
                col: row[col] for col in columns_to_save if col in row
            }
    
    def get_filtered_data(self) -> Optional[Dict[int, Dict[str, Any]]]:
        """Возвращает отфильтрованные данные."""
        return self.filtered_data
    
    @staticmethod
    def is_empty(val: Any) -> bool:
        """
        Проверяет, является ли значение пустым.
        """
        return (val is None or
                (isinstance(val, str) and val.strip().lower() in ("", "nan", "n/a", "none", "-", "null")) or
                (hasattr(pd, 'isna') and pd.isna(val)))
    
    def get_column_values(self, get_column: str, 
                          foc_mode: bool = False,
                          skip_condition: Optional[callable] = None) -> List[str]:
        """
        Извлекает уникальные значения из указанной колонки.
        """
        if get_column not in self.columns_save:
            messagebox.showerror(
                "Ошибка", 
                f"Название столбцов не совпадают.\n{get_column} в {self.columns_save}"
            )
            return []
        
        result = []
        all_data_dict = self.get_dict_all_data() if foc_mode else None
        
        for key in self.filtered_data:
            # Применяем FOC-фильтрацию если нужно
            if foc_mode and skip_condition is not None:
                row_data = all_data_dict.get(key, {})
                if skip_condition(row_data):
                    continue
            
            value = self.filtered_data[key].get(get_column)
            if value is None or self.is_empty(value):
                continue
            
            value_str = str(value).strip()
            if not value_str:
                continue
            
            # Разбиваем по разделителям
            items = value_str.replace(", ", "|").split("|")
            
            for item in items:
                item = item.strip()
                if not item:
                    continue
                
                # Убираем дублирование если есть и ":" и "-"
                if ":" in item and "-" in item:
                    item = item[:len(item) // 2 + 1]
                
                if item not in result:
                    result.append(item)
        
        result.sort(reverse=True)
        result.append("")  # Пустая строка в конце (оригинальное поведение)
        return result
    
    def return_data(self) -> Optional[pd.DataFrame]:
        """Возвращает загруженные данные."""
        return self.data
    
class MultiSheetReader:
    """
    Класс для чтения нескольких листов из одного Excel-файла.
    """
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.sheets: Dict[str, Optional[pd.DataFrame]] = {}
    
    def load_sheets(self, sheet_names: List[str]) -> Dict[str, Optional[pd.DataFrame]]:
        """
        Загружает указанные листы из Excel-файла.
        """
        for sheet_name in sheet_names:
            try:
                data = pd.read_excel(self.file_path, sheet_name=sheet_name)
                self.sheets[sheet_name] = data
            except Exception as e:
                print(f"Ошибка при загрузке листа {sheet_name}: {e}")
                self.sheets[sheet_name] = None
        
        return self.sheets
    
    def get_sheet_as_dict(self, sheet_name: str) -> dict[Hashable, dict] | None:
        """
        Возвращает данные листа в виде словаря.
        
        :param sheet_name: имя листа
        :return: словарь {индекс: {колонка: значение}} или None
        """
        data = self.sheets.get(sheet_name)
        if data is None:
            return None
        
        return {index: row.to_dict() for index, row in data.iterrows()}
    
    def get_sheet(self, sheet_name: str) -> Optional[pd.DataFrame]:
        """Возвращает DataFrame указанного листа."""
        return self.sheets.get(sheet_name)
    
