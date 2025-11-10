import os
from datetime import datetime as dt

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

import JsonWork


def auto_fit_columns(sheet):
    for column_cells in sheet.columns:
        koo = JsonWork.JsonConfig()
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        adjusted_width = (max_length + koo.getConfigsimbaProgram())  # Немного увеличим для красоты
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width


class ExcelWriter:
    def __init__(self, file_path, min_prog: str = None):
        """
        Инициализация класса для работы с Excel файлом

        :param file_path: Полный путь к Excel файлу
        """
        self.file_path = file_path
        self.min_prog = min_prog

        self.fill_color1 = PatternFill(start_color="6f747c", fill_type="solid")
        self.fill_color2 = PatternFill(start_color="aaadb2", fill_type="solid")
        self.fill_color3 = PatternFill(start_color="d9d9d9", fill_type="solid")
        self.fill_color4 = PatternFill(start_color="808080", fill_type="solid")
        self.fill_color5 = PatternFill(start_color="D7E4BC", fill_type="solid")

        self.thin_border = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))

        self.alfavit = ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA',
                        'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL']

        self.config = JsonWork.JsonConfig()

    def write_to_sheet(self, data, sheet_name, start_row=1, start_col=1):
        """
        Запись данных в указанный лист Excel файла

        :param data: Данные для записи (список списков или DataFrame)
        :param sheet_name: Имя листа
        :param start_row: Начальная строка (по умолчанию 1)
        :param start_col: Начальный столбец (по умолчанию 1)
        """
        try:
            # Если файл существует, загружаем его
            if os.path.exists(self.file_path):
                # Загружаем существующую книгу
                book = load_workbook(self.file_path)

                # Если лист существует, удаляем его
                if sheet_name in book.sheetnames:
                    del book[sheet_name]

                # Создаем новый лист
                sheet = book.create_sheet(sheet_name)

                # Записываем данные
                if isinstance(data, pd.DataFrame):
                    # Если данные в формате DataFrame
                    for r_idx, row in enumerate(data.values, start=start_row):
                        for c_idx, value in enumerate(row, start=start_col):
                            sheet.cell(row=r_idx, column=c_idx, value=value)
                else:
                    # Если данные в формате списка списков
                    print("self.min_prog: ", self.min_prog)
                    if self.min_prog == "BAM":
                        row_pack_h = 1
                        row_pack_d = 1
                        for r_idx, row in enumerate(data, start=start_row):
                            for c_idx, value in enumerate(row, start=start_col):
                                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                                if c_idx == 1 and value != "" and value != "/../":
                                    row_pack_h = r_idx
                                if c_idx == 1 and value == "/../":
                                    row_pack_d = r_idx
                                    cell = sheet.cell(row=r_idx, column=c_idx, value="")
                                if c_idx == 1:
                                    cell.fill = self.fill_color3
                                    cell.font = Font(color="f2ecde")
                                if r_idx == row_pack_h:
                                    cell.fill = self.fill_color1
                                    cell.font = Font(color="f2ecde")
                                    if c_idx == 12:
                                        cell.font = Font(color="6f747c")
                                elif r_idx == row_pack_d:
                                    cell.fill = self.fill_color2
                                    cell.font = Font(color="000000")
                                    if c_idx == 12:
                                        cell.font = Font(color="aaadb2")
                        auto_fit_columns(sheet)
                        sheet.column_dimensions['A'].width = 8
                        sheet.column_dimensions['B'].width = 26
                        sheet.column_dimensions['D'].width = 32
                        sheet.column_dimensions['E'].width = 35
                    elif self.min_prog == "Jp":
                        for r_idx, row in enumerate(data, start=start_row):
                            for c_idx, value in enumerate(row, start=start_col):
                                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                                if c_idx == 2 and value != "":
                                    cell.fill = self.fill_color3
                                    cell.font = Font(color="000000")
                                if (r_idx == 1 or r_idx == 5 or r_idx == 5 + 3 + int(self.config.getJPColumnName(
                                        "Table of contents: List_date"))) and value != "":
                                    cell.fill = self.fill_color1
                                    cell.font = Font(color="f2ecde")
                        auto_fit_columns(sheet)
                    elif self.min_prog == "Cz":
                        for r_idx, row in enumerate(data, start=start_row):
                            for c_idx, value in enumerate(row, start=start_col):
                                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                                if r_idx == 3 and value != "":
                                    cell.fill = self.fill_color3
                                    cell.font = Font(color="000000")
                                if r_idx == 1 or c_idx == 1:
                                    cell.fill = self.fill_color1
                                    cell.font = Font(color="f2ecde")
                        auto_fit_columns(sheet)
                    elif self.min_prog == "Ge":
                        for r_idx, row in enumerate(data, start=start_row):
                            for c_idx, value in enumerate(row, start=start_col):
                                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                                if r_idx in [9, 14, 19, 24, 29]:
                                    cell.fill = self.fill_color1
                                    cell.font = Font(color="f2ecde")
                                if r_idx in [6, 0] and value != "":
                                    cell.fill = self.fill_color1
                                    cell.font = Font(color="f2ecde")
                                if (c_idx in [3, 0] and value != "") or (r_idx == 5 and c_idx > 8):
                                    cell.fill = self.fill_color3
                                    cell.font = Font(color="000000")
                                if r_idx == 5 and 5 < c_idx < 9:
                                    cell.fill = self.fill_color4
                                    cell.font = Font(color="000000")
                                if 8 < c_idx < len(data[5]):
                                    if int(str(dt.now())[8:10]) == data[5][c_idx] and r_idx in [13, 10, 11, 12, 18, 15,
                                                                                                16, 17, 23, 20, 21, 22,
                                                                                                28, 25, 26, 27, 33, 30,
                                                                                                31, 32]:
                                        cell.fill = self.fill_color5
                                        cell.border = self.thin_border
                                        cell.font = Font(color="000000")

                                if value == r"\..../": cell = sheet.cell(row=r_idx, column=c_idx,
                                                                         value=f"={self.alfavit[c_idx - 9]}{r_idx - 2}")
                                if value == r"\.../": cell = sheet.cell(row=r_idx, column=c_idx,
                                                                        value=f"={self.alfavit[c_idx - 9]}{r_idx - 2}+{self.alfavit[c_idx - 10]}{r_idx}")
                        auto_fit_columns(sheet)
                        for i in self.alfavit:
                            sheet.column_dimensions[f'{i}'].width = 3.7
                        sheet.column_dimensions[f'f'].width = 3.7
                        sheet.column_dimensions[f'g'].width = 3.7
                        sheet.column_dimensions[f'h'].width = 3.7 + 7.0
                        sheet.column_dimensions[f'i'].width = 3.7 + 7.0
                    else:
                        for r_idx, row in enumerate(data, start=start_row):
                            for c_idx, value in enumerate(row, start=start_col):
                                sheet.cell(row=r_idx, column=c_idx, value=value)

                book.save(self.file_path)

            else:
                if isinstance(data, pd.DataFrame):
                    with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                        data.to_excel(writer, sheet_name=sheet_name, startrow=start_row - 1, startcol=start_col - 1,
                                      index=False)
                else:
                    df = pd.DataFrame(data)
                    with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, startrow=start_row - 1, startcol=start_col - 1,
                                    index=False)

            print(f"Данные успешно записаны в лист '{sheet_name}' файла {self.file_path}")

        except Exception as e:
            print(f"Ошибка при записи в Excel: {e}")

    def append_to_sheet(self, data, sheet_name):
        """
        Добавление данных в конец существующего листа

        :param data: Данные для добавления
        :param sheet_name: Имя листа
        """
        try:
            if os.path.exists(self.file_path):
                book = load_workbook(self.file_path)

                if sheet_name in book.sheetnames:
                    sheet = book[sheet_name]
                    # Определяем следующую строку
                    start_row = sheet.max_row + 1
                else:
                    sheet = book.create_sheet(sheet_name)
                    start_row = 1

                # Записываем данные
                if isinstance(data, pd.DataFrame):
                    for r_idx, row in enumerate(data.values, start=start_row):
                        for c_idx, value in enumerate(row, start=1):
                            sheet.cell(row=r_idx, column=c_idx, value=value)
                else:
                    for r_idx, row in enumerate(data, start=start_row):
                        for c_idx, value in enumerate(row, start=1):
                            sheet.cell(row=r_idx, column=c_idx, value=value)

                book.save(self.file_path)
                print(f"Данные успешно добавлены в лист '{sheet_name}' файла {self.file_path}")
            else:
                # Если файла нет, просто создаем его
                self.write_to_sheet(data, sheet_name)

        except Exception as e:
            print(f"Ошибка при добавлении данных в Excel: {e}")
